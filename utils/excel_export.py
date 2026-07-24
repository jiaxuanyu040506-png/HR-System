"""
excel_export.py

Builds downloadable Excel workbooks from attendance matrices. Kept
separate from attendance.py since this is purely about formatting an
Excel file, not attendance business logic.
"""
from __future__ import annotations

import io
import pandas as pd
from openpyxl.styles import PatternFill

REST_DAY_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")


def _grey_out_rest_day_columns(writer, sheet_name: str, rest_day_columns: list[int], num_rows: int):
    """Rest Day has no text symbol — colour its whole column grey instead,
    matching the on-screen Styler in pages/11_Attendance.py."""
    ws = writer.sheets[sheet_name]
    for day in rest_day_columns:
        col_idx = day + 1  # +1 because column A is the Employee name (index_label)
        for row_idx in range(1, num_rows + 2):  # +1 for header row, +1 for 1-indexing
            ws.cell(row=row_idx, column=col_idx).fill = REST_DAY_FILL


def attendance_month_to_excel(matrix: pd.DataFrame, year: int, month: int, rest_day_columns: list[int] = None) -> bytes:
    """One sheet, employees x days for a single month."""
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        sheet_name = f"{year}-{month:02d}"
        matrix.to_excel(writer, sheet_name=sheet_name, index_label="Employee")
        if rest_day_columns:
            _grey_out_rest_day_columns(writer, sheet_name, rest_day_columns, len(matrix))
        _write_legend(writer)
    return buffer.getvalue()


def attendance_year_to_excel(matrices: dict[int, pd.DataFrame], year: int,
                              rest_day_columns_by_month: dict[int, list[int]] = None) -> bytes:
    """One sheet per month (12 sheets total) in a single workbook."""
    import calendar as _calendar

    rest_day_columns_by_month = rest_day_columns_by_month or {}
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for month, matrix in matrices.items():
            sheet_name = _calendar.month_abbr[month] + f" {year}"
            if matrix.empty:
                pd.DataFrame({"Note": ["No employees found."]}).to_excel(writer, sheet_name=sheet_name, index=False)
            else:
                matrix.to_excel(writer, sheet_name=sheet_name, index_label="Employee")
                rest_days = rest_day_columns_by_month.get(month, [])
                if rest_days:
                    _grey_out_rest_day_columns(writer, sheet_name, rest_days, len(matrix))
        _write_legend(writer)
    return buffer.getvalue()


def _write_legend(writer):
    legend = pd.DataFrame({
        "Code": ["(blank)", "/", "L", "H", "X", "P", "(grey column)"],
        "Meaning": ["Present", "Absent", "Late", "Half Day", "On Leave", "Public Holiday", "Rest Day (weekend)"],
    })
    legend.to_excel(writer, sheet_name="Legend", index=False)
